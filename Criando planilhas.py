import openpyxl

#criar uma planilha(book)
book = openpyxl.Workbook()

#Como visualizar paginas existentes
print(book.sheetnames)

#Como criar uma página
book.create_sheet('teste')

#Como selecionar uma página
teste_page = book['teste']
teste_page.append(['num', 'num', 'num'])
teste_page.append(['um','dois','tres'])
teste_page.append(['quatro','cinco','seis'])
teste_page.append(['sete','oito','nove'])

#Salvar a planilha
book.save('testão.xlsx')